from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, colors

from cliente.models import Cliente
from django.http.response import JsonResponse, HttpResponse
from producto.models import Producto
from reporte.models import ReporteVenta, ReporteDetalleVenta



from django.http.response import JsonResponse, HttpResponse
from venta.models import Venta
from datetime import date, datetime
from reportlab.pdfgen import canvas

# TICKET PDF
def facturapdf(request,id):
    #
    venta = Venta.objects.get(id = id)

    # ficha = Ficha.objects.get(id=factura.ficha.id)
    #
    # numerodeficha = ficha.codigo
    # carrera = ficha.carrera
    # nombre = ficha.nombre
    # vencimiento = ficha.vencimiento
    # importe = ficha.mensualidad
    # print vencimiento

    # print request
    # print request.id

    response = HttpResponse(content_type='application/pdf')

    fecha = datetime.now()

    nombredelpdf = str(fecha.date())+str(fecha.time())+'.pdf'

    response['Content-Disposition'] = 'attachment; filename= Factura ' + nombredelpdf

    p = canvas.Canvas(response)



    for i in range(0,1):
        # p.drawString(400, 810, "1 Esto es un recibo. "+id)
        p.drawString(400, 810, "numerodeficha")
        p.drawString(485, 810, "carrera")
        p.drawString(15, 740, "nombre")
        # p.drawString(270, 740, carrera)
        # p.drawString(400, 755, nombre)
        # p.drawString(15, 705, numerodeficha)
        # p.drawString(100, 705, '1')
        # p.drawString(180, 705, str(vencimiento))
        # p.drawString(270, 705, str(importe))
        # p.drawString(400, 705, '1')
        # p.drawString(445, 705, str(vencimiento))
        # p.drawString(520, 705, str(importe))

        p.showPage()


    p.save()
    return response

def recibopdf(request,id):
    #
    venta = Venta.objects.get(id = id)

    # ficha = Ficha.objects.get(id=factura.ficha.id)
    #
    # numerodeficha = ficha.codigo
    # carrera = ficha.carrera
    # nombre = ficha.nombre
    # vencimiento = ficha.vencimiento
    # importe = ficha.mensualidad
    # print vencimiento

    # print request
    # print request.id

    response = HttpResponse(content_type='application/pdf')

    fecha = datetime.now()

    nombredelpdf = str(fecha.date())+str(fecha.time())+'.pdf'

    response['Content-Disposition'] = 'attachment; filename= Ticket ' + nombredelpdf

    p = canvas.Canvas(response)



    for i in range(0,1):
        # p.drawString(400, 810, "1 Esto es un recibo. "+id)
        p.drawString(400, 810, "numerodeficha")
        p.drawString(485, 810, "carrera")
        p.drawString(15, 740, "nombre")
        # p.drawString(270, 740, carrera)
        # p.drawString(400, 755, nombre)
        # p.drawString(15, 705, numerodeficha)
        # p.drawString(100, 705, '1')
        # p.drawString(180, 705, str(vencimiento))
        # p.drawString(270, 705, str(importe))
        # p.drawString(400, 705, '1')
        # p.drawString(445, 705, str(vencimiento))
        # p.drawString(520, 705, str(importe))

        p.showPage()


    p.save()
    return response




def ventasreportes01(request,dia1,mes1,ano1,dia2,mes2,ano2,username1,cliente1,elimina2):
    reportesaexcel = ReporteVenta.objects.all()
    print 'los parametros'
    print len(reportesaexcel)
    print 'dia1:'
    print dia1
    print
    print 'mes1'
    print mes1
    print
    print 'ano1'
    print ano1
    print
    print 'dia2'
    print dia2
    print
    print 'mes2'
    print mes2
    print
    print 'ano2'
    print ano2
    print
    print 'username1'
    print username1
    if(username1!='nadita'):
        username2 = User.objects.get(username=username1)
        print 'username1'
        print username2

    print
    print 'cliente1'
    print cliente1
    print
    print 'elimina2'
    print elimina2
    print

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'RESUMEN'



    # PARSEO LOS DATOS ######################################################################################

    sdia1 = dia1
    if (int(dia1) < 10 and int(dia1) != 0):
        sdia1 = '0' + str(dia1)

    smes1 = mes1
    if (int(mes1) < 10 and int(mes1) != 0):
        smes1 = '0' + str(mes1)

    sdia2 = dia2
    if (int(dia2) < 10 and int(dia2) != 0):
        sdia2 = '0' + str(dia2)

    smes2 = mes2
    if (int(mes2) < 10 and int(mes2) != 0):
        smes2 = '0' + str(mes2)

    print 'jojo'
    print sdia1
    print smes1
    print sdia2
    print smes2
    print 'jojo2'
    print int(sdia1)*2
    print 'jojo3'
    print sdia2 + 'ivan'

    if(sdia1!='0' and smes1!='0'):
        print 'primer filtro'
        # 2018 - 06 - 13
        # 19:53:29.021111 - 04

        desdeaux = ano1+'-'+smes1+'-'+sdia1+' 00:00:00'
        reportesaexcel = reportesaexcel.filter(fecha__gte=desdeaux)

    if (sdia2 != '0' and smes2 != '0'):
        print 'segundo filtro'
        hastaaux = ano2 + '-' + smes2 + '-' + sdia2 + ' 23:59:59'
        reportesaexcel = reportesaexcel.filter(fecha__lte=hastaaux)

    if( username1 != 'nadita' ):
        print 'segundo filtro------------------------------------------------------'
        print username1
        user1 = User.objects.get(username=username1).id
        print user1
        reportesaexcel = reportesaexcel.filter(user=user1)
        print reportesaexcel
        print 'segundo filtro------------------------------------------------------'


    if (cliente1 != 'nadita'):
        print 'tercer filtro'
        # cliente1 = str(cliente1)
        # cliente1 = cliente1[:-1]
        cli = Cliente.objects.get(numerodedocumento=cliente1)
        reportesaexcel = reportesaexcel.filter(cliente=cli.id)

    if (elimina2 == 'No_eliminadas'):
        print 'cuarto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=False)

    if (elimina2 == 'Eliminadas'):
        print 'quinto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=True)


    cantidaddeelemento = len(reportesaexcel)
    print cantidaddeelemento
    if(cantidaddeelemento==0):
        print 'kore no hay elementos'
    else:
        print 'tiene elemento'


    personas = Cliente.objects.all()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'RESUMEN'



    c = ws['A1']
    c.font = Font(size=12)
    ws.merge_cells('A1:D1')

    a1 = ws['A1']
    ft = Font(color=colors.BLUE)
    a1.font = ft


    ## APARTIR DE AQUI EMPIEZA LAS VENTA O VENTAS ELIMINADAS

    vn1 = 'no'
    ve1 = 'no'
    cvn = 0
    cve = 0
    filaaux = 1

    for r in reportesaexcel:
        if(r.eliminado==False):
            vn1 = 'si'
            cvn = cvn + 1
            print 'eliminados false'
        if(r.eliminado==True):
            ve1 = 'si'
            cve = cve + 1
            print 'eliminados true'

    if(len(reportesaexcel)!=0):
        filaaux = filaaux + 1
        if(vn1=='si'):
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=False)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A'+ str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B'+ str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'
                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft
                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id).filter(eliminado=False)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS'

            a1 = ws[ac]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)


            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)


        filaaux = filaaux + 1
        if (ve1 == 'si'):
            print 'ventas eliminadas'
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS ELIMINADAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=True)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A' + str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.RED)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.RED)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.RED)
                a1.font = ft


                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.RED)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS ELIMINADAS'

            a1 = ws[ac]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)

            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)

        if (vn1 == 'si'):
            # detalle de los productos vendidos
            print 'lista de productos vendidos'
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'LISTA DE PRODUCTOS VENDIDOS'

            a1 = ws[ac]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            x = 'D' + str(filaaux)
            ws.merge_cells(ac + ':' + x)
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'Producto'

            a1 = ws[ac]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            ps = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + ps)
            b = 'C' + str(filaaux)
            ws[b] = 'Cantidad'

            a1 = ws[b]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            ps = 'D' + str(filaaux)
            ws.merge_cells(b + ':' + ps)

            # aqui cargo los nombres de los productos vendidos en un vecto y entro la cantidad
            p1 = []
            p12 = []
            p2 = []
            vventass = reportesaexcel.filter(eliminado=False)
            for vent in vventass:
                dtvs = ReporteDetalleVenta.objects.filter(venta=vent.id)
                for ff in dtvs:
                    p1.append(ff.producto.id)
                    p12.append(ff.producto.id)
                    p2.append(ff.cantidad)

            print 'p1'
            print p1

            print 'p2'
            print p2
            # print p2[3]

            # pds = []
            abc = int(0)
            # pds.append(p1[0])
            for indi1 in range(abc, len(p1) - 1):
                for indi2 in range(indi1 + 1, len(p1)):
                    if(p1[indi1]!='' and p1[indi1]==p1[indi2]):
                        # pds.append(p1[indi2])
                        p1[indi2] = ''

            print p1
            p3 = []
            for indi1 in range(abc, len(p1)):
                if(p1[indi1]!=''):
                    print
                    p3.append(p1[indi1])
            print p3

            p4 = []

            for i in range(0, len(p3)):
                print p3[i]
                sp = 0
                for g in range(0, len(p12)):
                    if(p3[i]==p12[g]):
                        sp = sp + int(p2[g])
                p4.append(sp)
            print p4


            # aqui cargo los producto y sus catidades vendidas al excel

            for i in range(0, len(p3)):
                productoto = Producto.objects.get(pk=p3[i]).nombre
                filaaux = filaaux + 1
                ac = 'A' + str(filaaux)
                ws[ac] = productoto
                c = ws[ac]
                c.font = Font(size=10)
                ps = 'B' + str(filaaux)
                ws.merge_cells(ac + ':' + ps)
                b = 'C' + str(filaaux)
                ws[b] = p4[i]
                c = ws[b]
                c.font = Font(size=10)
                ps = 'D' + str(filaaux)
                ws.merge_cells(b + ':' + ps)



    cont = 4
    for persona in personas:
        # ws.cell(row=cont, column=2).value = persona.id
        # ws.cell(row=cont, column=3).value = persona.nombre
        # ws.cell(row=cont, column=4).value = persona.numerodedocumento
        # ws.cell(row=cont, column=5).value = persona.eliminado
        cont = cont + 1
    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    # contenido = "attachment; filename={0}".format(nombre_archivo)
    # response["Content-Disposition"] = contenido
    # fechareportename =
    # ahora = time.strftime("%c")
    # fechareportename = str(ahora)


    # hasta aqui ivan ########################################################################################

    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")

    # wb = Workbook()
    # ws = wb.active
    # ws['A1'] = 'RESUMEN'
    import datetime
    x = datetime.datetime.now()
    fechareportename = 'reporte-venta-'+str(x)
    print fechareportename
    fechareportename = fechareportename.replace(' ','-').replace(':','-').replace(':','-').replace(':','-')
    print fechareportename
    v = fechareportename.split('.')
    fechareportename = v[0]+'.xlsx'
    print fechareportename


    print ("Fecha y hora = %s" % x)

    response['Content-Disposition'] = '; filename={0}'.format(fechareportename)
    wb.save(response)

    print 'ivan1'





    print 'ivan2'

    return response



def ventasreportes02(request,dia1,mes1,ano1,dia2,mes2,ano2,username1,cliente1,elimina2):
    reportesaexcel = ReporteVenta.objects.all()
    print 'los parametros'
    print len(reportesaexcel)
    print 'dia1:'
    print dia1
    print
    print 'mes1'
    print mes1
    print
    print 'ano1'
    print ano1
    print
    print 'dia2'
    print dia2
    print
    print 'mes2'
    print mes2
    print
    print 'ano2'
    print ano2
    print
    print 'username1'
    print username1
    if(username1!='nadita'):
        username2 = User.objects.get(username=username1)
        print 'username1'
        print username2

    print
    print 'cliente1'
    print cliente1
    print
    print 'elimina2'
    print elimina2
    print

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'RESUMEN'



    # PARSEO LOS DATOS ######################################################################################

    sdia1 = dia1
    if (int(dia1) < 10 and int(dia1) != 0):
        sdia1 = '0' + str(dia1)

    smes1 = mes1
    if (int(mes1) < 10 and int(mes1) != 0):
        smes1 = '0' + str(mes1)

    sdia2 = dia2
    if (int(dia2) < 10 and int(dia2) != 0):
        sdia2 = '0' + str(dia2)

    smes2 = mes2
    if (int(mes2) < 10 and int(mes2) != 0):
        smes2 = '0' + str(mes2)

    print 'jojo'
    print sdia1
    print smes1
    print sdia2
    print smes2
    print 'jojo2'
    print int(sdia1)*2
    print 'jojo3'
    print sdia2 + 'ivan'

    if(sdia1!='0' and smes1!='0'):
        print 'primer filtro'
        # 2018 - 06 - 13
        # 19:53:29.021111 - 04

        desdeaux = ano1+'-'+smes1+'-'+sdia1+' 00:00:00'
        reportesaexcel = reportesaexcel.filter(fecha__gte=desdeaux)

    if (sdia2 != '0' and smes2 != '0'):
        print 'segundo filtro'
        hastaaux = ano2 + '-' + smes2 + '-' + sdia2 + ' 23:59:59'
        reportesaexcel = reportesaexcel.filter(fecha__lte=hastaaux)

    if( username1 != 'nadita' ):
        print 'segundo filtro------------------------------------------------------'
        print username1
        user1 = User.objects.get(username=username1).id
        print user1
        reportesaexcel = reportesaexcel.filter(user=user1)
        print reportesaexcel
        print 'segundo filtro------------------------------------------------------'


    if (cliente1 != 'nadita'):
        print 'tercer filtro'
        # cliente1 = str(cliente1)
        # cliente1 = cliente1[:-1]
        cli = Cliente.objects.get(numerodedocumento=cliente1)
        reportesaexcel = reportesaexcel.filter(cliente=cli.id)

    if (elimina2 == 'No_eliminadas'):
        print 'cuarto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=False)

    if (elimina2 == 'Eliminadas'):
        print 'quinto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=True)


    cantidaddeelemento = len(reportesaexcel)
    print cantidaddeelemento
    if(cantidaddeelemento==0):
        print 'kore no hay elementos'
    else:
        print 'tiene elemento'


    personas = Cliente.objects.all()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'LISTA DE PRODUCTOS VENDIDOS'



    c = ws['A1']
    c.font = Font(size=12)
    ws.merge_cells('A1:D1')

    a1 = ws['A1']
    ft = Font(color=colors.BLUE)
    a1.font = ft




    ## APARTIR DE AQUI EMPIEZA LAS VENTA O VENTAS ELIMINADAS

    vn1 = 'no'
    ve1 = 'no'
    cvn = 0
    cve = 0
    filaaux = 1

    for r in reportesaexcel:
        if(r.eliminado==False):
            vn1 = 'si'
            cvn = cvn + 1
            print 'eliminados false'
        if(r.eliminado==True):
            ve1 = 'si'
            cve = cve + 1
            print 'eliminados true'

    if(len(reportesaexcel)!=0):




        filaaux = filaaux + 1
        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Segun estos datos'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        filaaux3 = 'D' + str(filaaux)
        ws.merge_cells(filaaux2+':'+filaaux3)


        # Ventas--------------------------------
        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Ventas'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        # filaaux3 = 'B' + str(filaaux)
        # ws.merge_cells(filaaux2 + ':' + filaaux3)

        c = 'B' + str(filaaux)
        ws[c] = elimina2
        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)



        # Clientes-----------------------------

        # clientereporte01=cliente1
        if(cliente1 == 'nadita'):
            clientereporte01 = 'Todos'
        else:
            clientereporte01 = str(Cliente.objects.get(numerodedocumento=cliente1).nombre +' '+Cliente.objects.get(numerodedocumento=cliente1).numerodedocumento)

        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Cliente'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        # filaaux3 = 'B' + str(filaaux)
        # ws.merge_cells(filaaux2 + ':' + filaaux3)
        c = 'B' + str(filaaux)
        ws[c] = clientereporte01
        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)

        # cajero -----------------------------

        if (username1 == 'nadita'):
            username1reporte01 = 'Todos'
            print


        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Cajero'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        # filaaux3 = 'B' + str(filaaux)
        # ws.merge_cells(filaaux2 + ':' + filaaux3)
        c = 'B' + str(filaaux)
        ws[c] = username1reporte01
        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)


        # Fecha desde --------------------------------

        if(dia1 == '0' and mes1=='0' and ano1=='0'):
            fechadesdereporte01 = 'no especifico fecha'
            print
        else:
            fechadesdereporte01 = ano1 + '-' + smes1 + '-' + sdia1 + ' 00:00:00'

        # fechadesdereporte01 = desdeaux
        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Desde'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        # filaaux3 = 'B' + str(filaaux)
        # ws.merge_cells(filaaux2 + ':' + filaaux3)
        c = 'B' + str(filaaux)
        ws[c] = str(fechadesdereporte01)
        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)

        # Fecha hasta -------------------------------------

        if (dia2 == '0' and mes2 == '0' and ano2 == '0'):
            fechahastareporte01 = 'no especifico fecha'
            print
        else:
            fechahastareporte01 = ano2 + '-' + smes2 + '-' + sdia2 + ' 23:59:59'
        # fechahastareporte01 = hastaaux
        filaaux = filaaux + 1
        filaaux2 = 'A' + str(filaaux)
        ws[filaaux2] = 'Hasta'

        a1 = ws[filaaux2]
        ft = Font(color=colors.BLUE)
        a1.font = ft

        # filaaux3 = 'B' + str(filaaux)
        # ws.merge_cells(filaaux2 + ':' + filaaux3)
        c = 'B' + str(filaaux)
        ws[c] = str(fechahastareporte01)

        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)



        if(False):
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=False)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A'+ str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B'+ str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'
                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft
                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id).filter(eliminado=False)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS'

            a1 = ws[ac]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)


            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)


        # filaaux = filaaux + 1
        if (False):
            print 'ventas eliminadas'
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS ELIMINADAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=True)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A' + str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.RED)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.RED)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.RED)
                a1.font = ft


                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.RED)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS ELIMINADAS'

            a1 = ws[ac]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)

            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)

        if (vn1 == 'si'):
            # detalle de los productos vendidos
            print 'lista de productos vendidos'
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'LISTA DE PRODUCTOS VENDIDOS'

            a1 = ws[ac]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            x = 'D' + str(filaaux)
            ws.merge_cells(ac + ':' + x)
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'Producto'

            a1 = ws[ac]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            ps = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + ps)
            b = 'C' + str(filaaux)
            ws[b] = 'Cantidad'

            a1 = ws[b]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            ps = 'D' + str(filaaux)
            ws.merge_cells(b + ':' + ps)

            # aqui cargo los nombres de los productos vendidos en un vecto y entro la cantidad
            p1 = []
            p12 = []
            p2 = []
            vventass = reportesaexcel.filter(eliminado=False)
            for vent in vventass:
                dtvs = ReporteDetalleVenta.objects.filter(venta=vent.id).filter(eliminado=False)
                for ff in dtvs:
                    p1.append(ff.producto.id)
                    p12.append(ff.producto.id)
                    p2.append(ff.cantidad)

            print 'p1'
            print p1

            print 'p2'
            print p2
            # print p2[3]

            # pds = []
            abc = int(0)
            # pds.append(p1[0])
            for indi1 in range(abc, len(p1) - 1):
                for indi2 in range(indi1 + 1, len(p1)):
                    if(p1[indi1]!='' and p1[indi1]==p1[indi2]):
                        # pds.append(p1[indi2])
                        p1[indi2] = ''

            print p1
            p3 = []
            for indi1 in range(abc, len(p1)):
                if(p1[indi1]!=''):
                    print
                    p3.append(p1[indi1])
            print p3

            p4 = []

            for i in range(0, len(p3)):
                print p3[i]
                sp = 0
                for g in range(0, len(p12)):
                    if(p3[i]==p12[g]):
                        sp = sp + int(p2[g])
                p4.append(sp)
            print p4


            # aqui cargo los producto y sus catidades vendidas al excel

            for i in range(0, len(p3)):
                productoto = Producto.objects.get(pk=p3[i]).nombre
                filaaux = filaaux + 1
                ac = 'A' + str(filaaux)
                ws[ac] = productoto
                c = ws[ac]
                c.font = Font(size=10)
                ps = 'B' + str(filaaux)
                ws.merge_cells(ac + ':' + ps)
                b = 'C' + str(filaaux)
                ws[b] = p4[i]
                c = ws[b]
                c.font = Font(size=10)
                ps = 'D' + str(filaaux)
                ws.merge_cells(b + ':' + ps)



    cont = 4
    for persona in personas:
        # ws.cell(row=cont, column=2).value = persona.id
        # ws.cell(row=cont, column=3).value = persona.nombre
        # ws.cell(row=cont, column=4).value = persona.numerodedocumento
        # ws.cell(row=cont, column=5).value = persona.eliminado
        cont = cont + 1
    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    # contenido = "attachment; filename={0}".format(nombre_archivo)
    # response["Content-Disposition"] = contenido
    # fechareportename =
    # ahora = time.strftime("%c")
    # fechareportename = str(ahora)


    # hasta aqui ivan ########################################################################################

    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")

    # wb = Workbook()
    # ws = wb.active
    # ws['A1'] = 'RESUMEN'
    import datetime
    x = datetime.datetime.now()
    fechareportename = 'productos-vendidos-'+str(x)
    print fechareportename
    fechareportename = fechareportename.replace(' ','-').replace(':','-').replace(':','-').replace(':','-')
    print fechareportename
    v = fechareportename.split('.')
    fechareportename = v[0]+'.xlsx'
    print fechareportename


    print ("Fecha y hora = %s" % x)

    response['Content-Disposition'] = '; filename={0}'.format(fechareportename)
    wb.save(response)

    print 'ivan1'





    print 'ivan2'

    return response


def ventasreportes03(request,dia1,mes1,ano1,dia2,mes2,ano2,username1,cliente1,elimina2):
    reportesaexcel = ReporteVenta.objects.all()
    print 'los parametros'
    print len(reportesaexcel)
    print 'dia1:'
    print dia1
    print
    print 'mes1'
    print mes1
    print
    print 'ano1'
    print ano1
    print
    print 'dia2'
    print dia2
    print
    print 'mes2'
    print mes2
    print
    print 'ano2'
    print ano2
    print
    print 'username1'
    print username1
    if(username1!='nadita'):
        username2 = User.objects.get(username=username1)
        print 'username1'
        print username2

    print
    print 'cliente1'
    print cliente1
    print
    print 'elimina2'
    print elimina2
    print

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'RESUMEN'



    # PARSEO LOS DATOS ######################################################################################

    sdia1 = dia1
    if (int(dia1) < 10 and int(dia1) != 0):
        sdia1 = '0' + str(dia1)

    smes1 = mes1
    if (int(mes1) < 10 and int(mes1) != 0):
        smes1 = '0' + str(mes1)

    sdia2 = dia2
    if (int(dia2) < 10 and int(dia2) != 0):
        sdia2 = '0' + str(dia2)

    smes2 = mes2
    if (int(mes2) < 10 and int(mes2) != 0):
        smes2 = '0' + str(mes2)

    print 'jojo'
    print sdia1
    print smes1
    print sdia2
    print smes2
    print 'jojo2'
    print int(sdia1)*2
    print 'jojo3'
    print sdia2 + 'ivan'

    if(sdia1!='0' and smes1!='0'):
        print 'primer filtro'
        # 2018 - 06 - 13
        # 19:53:29.021111 - 04

        desdeaux = ano1+'-'+smes1+'-'+sdia1+' 00:00:00'
        reportesaexcel = reportesaexcel.filter(fecha__gte=desdeaux)

    if (sdia2 != '0' and smes2 != '0'):
        print 'segundo filtro'
        hastaaux = ano2 + '-' + smes2 + '-' + sdia2 + ' 23:59:59'
        reportesaexcel = reportesaexcel.filter(fecha__lte=hastaaux)

    if( username1 != 'nadita' ):
        print 'segundo filtro------------------------------------------------------'
        print username1
        user1 = User.objects.get(username=username1).id
        print user1
        reportesaexcel = reportesaexcel.filter(user=user1)
        print reportesaexcel
        print 'segundo filtro------------------------------------------------------'


    if (cliente1 != 'nadita'):
        print 'tercer filtro'
        # cliente1 = str(cliente1)
        # cliente1 = cliente1[:-1]
        cli = Cliente.objects.get(numerodedocumento=cliente1)
        reportesaexcel = reportesaexcel.filter(cliente=cli.id)

    if (elimina2 == 'No_eliminadas'):
        print 'cuarto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=False)

    if (elimina2 == 'Eliminadas'):
        print 'quinto filtro'
        reportesaexcel = reportesaexcel.filter(eliminado=True)


    cantidaddeelemento = len(reportesaexcel)
    print cantidaddeelemento
    if(cantidaddeelemento==0):
        print 'kore no hay elementos'
    else:
        print 'tiene elemento'


    personas = Cliente.objects.all()

    filaaux = 0


    import datetime
    x = datetime.datetime.now()
    fechareportename = 'stock-producto-' + str(x)
    print fechareportename
    fechareportename = fechareportename.replace(' ', '-').replace(':', '-').replace(':', '-').replace(':', '-')
    print fechareportename
    v = fechareportename.split('.')
    fechareportename = v[0] + '.xlsx'
    print fechareportename




    wb = Workbook()
    ws = wb.active



    filaaux = filaaux + 1
    a = 'A' + str(filaaux)
    ws[a] = 'STOCK DE PRODUCTOS'

    c = ws[a]
    c.font = Font(size=12)
    d = 'D' + str(filaaux)
    ws.merge_cells(a+':'+d)
    a1 = ws[a]
    ft = Font(color=colors.BLUE)
    a1.font = ft



    filaaux = filaaux + 1
    a = 'A' + str(filaaux)
    ws[a] = 'Fecha'

    a1 = ws[a]
    ft = Font(color=colors.BLUE)
    a1.font = ft

    b = 'B' + str(filaaux)
    auxf = []
    auxf = str(x).split('.')
    ws[b] = str(auxf[0])
    d = 'D' + str(filaaux)
    ws.merge_cells(b+':'+d)






    filaaux = filaaux + 1
    a = 'A' + str(filaaux)
    ws[a] = 'Productos'
    a1 = ws[a]
    ft = Font(color=colors.BLUE)
    a1.font = ft
    b = 'B' + str(filaaux)
    ws.merge_cells(a + ':' + b)

    c = 'C' + str(filaaux)
    ws[c] = 'Cantidad'
    a1 = ws[c]
    ft = Font(color=colors.BLUE)
    a1.font = ft
    d = 'D' + str(filaaux)
    ws.merge_cells(c + ':' + d)

    productosreporte01 = Producto.objects.all()

    for i in productosreporte01:
        filaaux = filaaux + 1
        a = 'A' + str(filaaux)
        ws[a] = i.nombre
        b = 'B' + str(filaaux)
        ws.merge_cells(a + ':' + b)

        c = 'C' + str(filaaux)
        ws[c] = i.cantidadenexistencia
        d = 'D' + str(filaaux)
        ws.merge_cells(c + ':' + d)


    ## APARTIR DE AQUI EMPIEZA LAS VENTA O VENTAS ELIMINADAS

    vn1 = 'no'
    ve1 = 'no'
    cvn = 0
    cve = 0


    for r in reportesaexcel:
        if(r.eliminado==False):
            vn1 = 'si'
            cvn = cvn + 1
            print 'eliminados false'
        if(r.eliminado==True):
            ve1 = 'si'
            cve = cve + 1
            print 'eliminados true'

    if(len(reportesaexcel)!=0):
        filaaux = filaaux + 1
        if(False):
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=False)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A'+ str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B'+ str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'
                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft
                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id).filter(eliminado=False)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.BLUE)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS'

            a1 = ws[ac]
            ft = Font(color=colors.BLUE)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)


            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)


        filaaux = filaaux + 1
        if (False):
            print 'ventas eliminadas'
            filaaux = filaaux + 1
            filaaux2 = 'A' + str(filaaux)
            ws[filaaux2] = 'VENTAS ELIMINADAS'

            a1 = ws[filaaux2]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'D' + str(filaaux)
            ws.merge_cells(filaaux2 + ':' + d)
            objetosreportes = reportesaexcel.filter(eliminado=True)
            sumaventastotal1 = 0
            for i in objetosreportes:

                # fecha
                filaaux = filaaux + 2
                af = 'A' + str(filaaux)
                ws[af] = 'Fecha:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                fechafecha = str(i.fecha.today())
                fechafecha2 = fechafecha.split('.')
                ws[bf] = str(fechafecha2[0])
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # nombre
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nombre:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                nombrenombre = str(i.cliente.nombre)
                ws[bf] = str(nombrenombre)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # numero de documento
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Nro. Doc.:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                numerod = str(i.cliente.numerodedocumento)
                ws[bf] = str(numerod)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # usuario
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Usuario:'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                usuarioexcel = str(i.user.username)
                ws[bf] = str(usuarioexcel)
                c = ws[bf]
                c.font = Font(size=10)
                d = 'D' + str(filaaux)
                ws.merge_cells(bf + ':' + d)

                # detalle venta
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Detalle Venta'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                # bf = 'B' + str(filaaux)
                # usuarioexcel = str(i.user.username)
                # ws[bf] = str(usuarioexcel)
                d = 'D' + str(filaaux)
                ws.merge_cells(af + ':' + d)

                # titulos: producto precio cantidad subtotal
                # producto
                filaaux = filaaux + 1
                af = 'A' + str(filaaux)
                ws[af] = 'Producto'

                a1 = ws[af]
                ft = Font(color=colors.RED)
                a1.font = ft

                bf = 'B' + str(filaaux)
                ws[bf] = 'Precio de venta'

                a1 = ws[bf]
                ft = Font(color=colors.RED)
                a1.font = ft

                cf = 'C' + str(filaaux)
                ws[cf] = 'cantidad'

                a1 = ws[cf]
                ft = Font(color=colors.RED)
                a1.font = ft

                df = 'D' + str(filaaux)
                ws[df] = 'subTotal'

                a1 = ws[df]
                ft = Font(color=colors.RED)
                a1.font = ft


                # detalle reporte venta cargar
                rdv = ReporteDetalleVenta.objects.filter(venta=i.id)
                sumasut = 0
                for e in rdv:
                    print
                    filaaux = filaaux + 1
                    af = 'A' + str(filaaux)
                    ws[af] = e.producto.nombre
                    c = ws[af]
                    c.font = Font(size=10)
                    bf = 'B' + str(filaaux)
                    ws[bf] = e.preciounitario
                    c = ws[bf]
                    c.font = Font(size=10)
                    cf = 'C' + str(filaaux)
                    ws[cf] = e.cantidad
                    c = ws[cf]
                    c.font = Font(size=10)
                    df = 'D' + str(filaaux)
                    ws[df] = e.subtotal
                    c = ws[df]
                    c.font = Font(size=10)
                    sumasut = sumasut + int(e.subtotal)

                # TOTAL DE VA VENTA 1
                filaaux = filaaux + 1
                ac = 'C' + str(filaaux)
                ws[ac] = 'Total'

                a1 = ws[ac]
                ft = Font(color=colors.RED)
                a1.font = ft

                ad = 'D' + str(filaaux)
                ws[ad] = sumasut
                c = ws[ad]
                c.font = Font(size=10)
                sumaventastotal1 = sumaventastotal1 + sumasut
                # colocamos la suma de todas las ventas
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'TOTAL DE VENTAS ELIMINADAS'

            a1 = ws[ac]
            ft = Font(color=colors.RED)
            a1.font = ft

            d = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + d)

            cd = 'C' + str(filaaux)
            ws[cd] = sumaventastotal1
            c = ws[cd]
            c.font = Font(size=10)
            x = 'D' + str(filaaux)
            ws.merge_cells(cd + ':' + x)

        if (False):
            # detalle de los productos vendidos
            print 'lista de productos vendidos'
            filaaux = filaaux + 1
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'LISTA DE PRODUCTOS VENDIDOS'

            a1 = ws[ac]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            x = 'D' + str(filaaux)
            ws.merge_cells(ac + ':' + x)
            filaaux = filaaux + 1
            ac = 'A' + str(filaaux)
            ws[ac] = 'Producto'

            a1 = ws[ac]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            ps = 'B' + str(filaaux)
            ws.merge_cells(ac + ':' + ps)
            b = 'C' + str(filaaux)
            ws[b] = 'Cantidad'

            a1 = ws[b]
            ft = Font(color=colors.GREEN)
            a1.font = ft

            ps = 'D' + str(filaaux)
            ws.merge_cells(b + ':' + ps)

            # aqui cargo los nombres de los productos vendidos en un vecto y entro la cantidad
            p1 = []
            p12 = []
            p2 = []
            vventass = reportesaexcel.filter(eliminado=False)
            for vent in vventass:
                dtvs = ReporteDetalleVenta.objects.filter(venta=vent.id)
                for ff in dtvs:
                    p1.append(ff.producto.id)
                    p12.append(ff.producto.id)
                    p2.append(ff.cantidad)

            print 'p1'
            print p1

            print 'p2'
            print p2
            # print p2[3]

            # pds = []
            abc = int(0)
            # pds.append(p1[0])
            for indi1 in range(abc, len(p1) - 1):
                for indi2 in range(indi1 + 1, len(p1)):
                    if(p1[indi1]!='' and p1[indi1]==p1[indi2]):
                        # pds.append(p1[indi2])
                        p1[indi2] = ''

            print p1
            p3 = []
            for indi1 in range(abc, len(p1)):
                if(p1[indi1]!=''):
                    print
                    p3.append(p1[indi1])
            print p3

            p4 = []

            for i in range(0, len(p3)):
                print p3[i]
                sp = 0
                for g in range(0, len(p12)):
                    if(p3[i]==p12[g]):
                        sp = sp + int(p2[g])
                p4.append(sp)
            print p4


            # aqui cargo los producto y sus catidades vendidas al excel

            for i in range(0, len(p3)):
                productoto = Producto.objects.get(pk=p3[i]).nombre
                filaaux = filaaux + 1
                ac = 'A' + str(filaaux)
                ws[ac] = productoto
                c = ws[ac]
                c.font = Font(size=10)
                ps = 'B' + str(filaaux)
                ws.merge_cells(ac + ':' + ps)
                b = 'C' + str(filaaux)
                ws[b] = p4[i]
                c = ws[b]
                c.font = Font(size=10)
                ps = 'D' + str(filaaux)
                ws.merge_cells(b + ':' + ps)



    cont = 4
    for persona in personas:
        # ws.cell(row=cont, column=2).value = persona.id
        # ws.cell(row=cont, column=3).value = persona.nombre
        # ws.cell(row=cont, column=4).value = persona.numerodedocumento
        # ws.cell(row=cont, column=5).value = persona.eliminado
        cont = cont + 1
    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    # contenido = "attachment; filename={0}".format(nombre_archivo)
    # response["Content-Disposition"] = contenido
    # fechareportename =
    # ahora = time.strftime("%c")
    # fechareportename = str(ahora)


    # hasta aqui ivan ########################################################################################

    nombre_archivo = "ReportePersonasExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")

    # wb = Workbook()
    # ws = wb.active
    # ws['A1'] = 'RESUMEN'



    print ("Fecha y hora = %s" % x)

    response['Content-Disposition'] = '; filename={0}'.format(fechareportename)
    wb.save(response)

    print 'ivan1'





    print 'ivan2'

    return response